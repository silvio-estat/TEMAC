"""
PRISMA Scoring — PPCOT / Data Lake C2
======================================
Implementa a triagem bibliométrica automatizada dos 586 artigos do corpus C2
(ICCRTS 2016–2026, C2COE, MOR) para a dissertação:

    "O Combustível da Decisão: Uma Arquitetura de Data Lake para Acelerar
     o Processo de Planejamento e Condução das Operações Terrestres (PPCOT)"

Uso:
    python3 prisma_scoring.py

Entradas:
    XLSX_IN  — planilha original do corpus (título + resumo por artigo)
    CSV_IN   — versão CSV gerada automaticamente via LibreOffice (se ausente)

Saídas:
    JSON_OUT — ranking completo, usado por outros scripts
    XLSX_OUT — planilha de auditoria com três abas:
               • "Candidatos_33"   : lista final de 33 artigos
               • "Scoring_Detalhe": pontos por categoria para cada artigo
               • "Metodologia"    : parâmetros, pesos e justificativas

──────────────────────────────────────────────────────────────────────────────
METODOLOGIA E JUSTIFICATIVA DAS DECISÕES
──────────────────────────────────────────────────────────────────────────────

PASSO 1 — CATEGORIAS E PESOS
─────────────────────────────
Foram definidas cinco categorias de termos. O PROBLEMA CENTRAL DA DISSERTAÇÃO
é a ausência de uma infraestrutura de dados unificada que conecte o PITCIC ao
ciclo de decisão C2. Por isso as categorias de dados (A, B, E) recebem pesos
maiores que a categoria de ML/IA (D):

  A — Arquitetura de dados
      Fundamento: a dissertação propõe um artefato tecnológico (Data Lake/
      Lakehouse). Artigos que descrevem componentes análogos (data lake,
      big data, data pipeline, data management etc.) contribuem diretamente
      para a fundamentação arquitetural do artefato.
      → Sub-categoria A_HIGH (termos específicos): peso 6/título, 3/abstract
      → Sub-categoria A_LOW  (termos genéricos)  : peso 3/título, 1/abstract
      A divisão em HIGH/LOW existe porque termos como "architecture" sozinho
      capturam artigos de arquitetura de comunicações ou sistemas que não
      tratam de dados; por isso seu peso é menor.

  B — Integração de dados heterogêneos
      Fundamento: o PITCIC agrega fontes heterogêneas (HUMINT, SIGINT, OSINT,
      imagem, relatórios textuais). Artigos que discutem integração, fusão ou
      agregação de dados distribuídos e heterogêneos definem os REQUISITOS DE
      INGESTÃO do Data Lake proposto.
      → Sub-categoria B_HIGH (termos específicos): peso 5/título, 3/abstract
      → Sub-categoria B_LOW  (termos genéricos)  : peso 3/título, 1/abstract

  D — ML / IA em C2
      Fundamento: ML/IA é condição habilitante para o uso efetivo dos dados
      do Data Lake no ciclo decisório. No entanto, o corpus inteiro tende a
      mencionar "AI" e "machine learning" — a densidade desses termos não
      diferencia artigos de arquitetura de dados de artigos de outra natureza.
      → Peso: 3/título, 2/abstract
      → CAP DE 6 PONTOS POR ARTIGO: impede que artigos puramente sobre AI
        genérica dominem o ranking apenas por acumular muitas ocorrências dos
        termos. Sem o cap, artigos sobre realidade aumentada, detecção de
        influência em redes sociais ou automação de UxVs atingem scores
        maiores que artigos sobre arquitetura de dados para C2, distorcendo
        a triagem. O valor 6 foi escolhido porque equivale a dois termos D
        no título — presença substantiva de AI sem permitir inflação por
        repetição.

  E — Governança e qualidade de dados
      Fundamento: para que a IA produza decisões confiáveis, os dados do
      Data Lake precisam de proveniência, qualidade e metadados rastreáveis.
      Artigos sobre governança definem os REQUISITOS DE CONFIABILIDADE do
      artefato.
      → Peso: 5/título, 2/abstract

  C — C2 / Decisão (CORTE MÍNIMO — não soma pontos)
      Fundamento: o corpus é composto inteiramente de artigos de C2, mas
      inclui também documentos institucionais, relatórios de exercício e
      textos doutrinários sem orientação técnica. O corte C garante que
      apenas artigos com conexão explícita a C2, decisão ou operações
      militares entrem no pool de candidatos. Como todos os artigos
      relevantes passam esse filtro, ele funciona como PORTÃO, não como
      diferenciador de relevância — por isso não soma pontos.

PASSO 2 — SCORE COMPOSTO COM BÔNUS ARQUITETURAL
─────────────────────────────────────────────────
Score_comp = A + B + D(com cap) + E  +  bônus(3) se (A + B) ≥ 3

O BÔNUS existe porque artigos que combinam relevância arquitetural (A+B ≥ 3)
com quaisquer outros critérios são precisamente o tipo de artigo mais valioso
para a dissertação. O bônus de 3 pontos eleva esses artigos acima de artigos
que pontuam exclusivamente em D, sem introduzir novos termos de busca.

Por que A+B ≥ 3 como threshold do bônus?
    - A_HIGH no abstract = 3 pontos: um único termo específico de arquitetura
      de dados no resumo já qualifica (ex: "data lake" no abstract)
    - A_LOW no título = 3 pontos: um único termo arquitetural no título qualifica
    - B_HIGH no abstract = 3 pontos: um único termo de integração no resumo qualifica
    O valor 3 é o menor peso que indica PRESENÇA REAL do conceito, não coincidência.

PASSO 3 — REGRA DE CORTE AUTOMÁTICO
──────────────────────────────────────
    score_comp ≥ 7  E  (A + B ≥ 3  OU  E ≥ 5)

Por que score_comp ≥ 7?
    Abaixo de 7, os artigos atingem o score exclusivamente via D=6 (cap) sem
    qualquer contribuição de A, B ou E. Esses artigos tratam de AI/ML em C2
    de forma genérica, sem abordar infraestrutura de dados — fora do escopo
    da dissertação.

Por que a condição (A+B ≥ 3 OU E ≥ 5)?
    O score_comp ≥ 7 não é suficiente por si só: artigos que atingem
    score_comp = 7 exclusivamente via D=6 + A_LOW(1) no abstract (1 ponto)
    ainda são artigos de AI genérica que mencionaram "architecture" ou
    "information system" acidentalmente. A condição adicional garante que:
      • A+B ≥ 3: o artigo aborda dados com profundidade mínima verificável
        (pelo menos um termo de arquitetura de dados no título ou no abstract)
      • OU E ≥ 5: o artigo aborda governança com peso de título (um termo de
        qualidade/proveniência/metadados no título do artigo)
    Essa conjunção (E não simples OR) evita que o corte seja driblado por
    artigos que atingem score_comp = 7 apenas acumulando termos de AI.

    Resultado: 29 artigos selecionados automaticamente.

PASSO 4 — ADIÇÕES MANUAIS (4 artigos)
────────────────────────────────────────
O PRISMA prevê que a triagem automática seja complementada pelo julgamento
do pesquisador na etapa de ELEGIBILIDADE. Quatro artigos foram adicionados
manualmente por relevância temática justificada, mesmo não atingindo o corte
automático:

    1. Schubert et al. (2018) — "AI for Decision Support in C2 Systems"
       Justificativa: único artigo do corpus com dados empíricos de workshop
       com militares sobre O QUE A IA deve fazer com os dados de C2 (funções
       S2/S3, COP, análise de ameaças, avaliação de COAs). Define os
       REQUISITOS FUNCIONAIS que o Data Lake deve atender. Score baixo no
       critério automático porque discute funções de IA, não infraestrutura.

    2. Abdelazez, Garber, Ghanmi (2023) — "Decision Support: AI for Wargaming"
       Justificativa: único artigo do corpus que cita explicitamente o
       MINISTÉRIO DA DEFESA DO BRASIL como planejando incorporar IA ao
       wargaming de COAs — conexão direta com o contexto nacional da
       dissertação. Aborda o Jogo da Guerra (Step 4 do PPM), principal
       CASO DE USO que o Data Lake deve habilitar.

    3. Firth, Poole, Turner (2024) — "Opportunity Knocks"
       Justificativa: do projeto MSC2/Dstl (UK). Argumenta empiricamente que
       a IA no planejamento operacional requer "data analytics" como condição
       necessária — VALIDAÇÃO EMPÍRICA do problema de pesquisa desta
       dissertação.

    4. Turner, Leggatt et al. (2024) — "A challenging case for AI"
       Justificativa: mesmo projeto MSC2. Documenta que a INDISPONIBILIDADE
       DE DADOS foi identificada por planejadores militares reais como a
       PRINCIPAL BARREIRA à IA no planejamento operacional — evidência
       empírica direta do problema que o Data Lake resolve.

    Nota metodológica: essas adições são documentadas explicitamente no
    relatório PRISMA (seção "Elegibilidade") para garantir transparência e
    rastreabilidade conforme exigido pela metodologia DSR (Hevner et al., 2004;
    Dresch et al., 2015).
"""

import csv
import re
import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuração de arquivos ──────────────────────────────────────────────────
XLSX_IN  = '/home/silvio/Projetos/TEMAC/ajustes_BD/BASES_C2_JUNTAS/bases_c2_juntas.xlsx'
CSV_IN   = '/tmp/bases_c2_juntas.csv'
JSON_OUT = '/tmp/scoring_result.json'
XLSX_OUT = '/home/silvio/Projetos/TEMAC/escolha_artigos_C2/scoring_auditoria.xlsx'

D_CAP       = 6    # pontuação máxima da categoria D por artigo
BONUS_AB    = 3    # bônus concedido quando A+B >= BONUS_THRESHOLD
BONUS_THRESHOLD = 3
COMP_MIN    = 7    # score_comp mínimo para entrada automática
AB_MIN      = 3    # mínimo de A+B para aprovação pelo corte
E_MIN       = 5    # alternativa: mínimo de E para aprovação pelo corte

# ── Termos por categoria ──────────────────────────────────────────────────────
CAT_A_HIGH = [
    'data lake', 'lakehouse', 'data warehouse', 'data architecture',
    'data infrastructure', 'data platform', 'data repository', 'data store',
    'big data', 'data pipeline', 'data management', 'information architecture',
    'data science', 'ETL',
]
CAT_A_LOW = [
    'architecture', 'architectural', 'information system', 'knowledge base',
    'service-oriented',
]
CAT_B_HIGH = [
    'heterogeneous data', 'data fusion', 'data integration', 'information fusion',
    'information integration', 'multi-source', 'multi-int', 'sensor fusion',
    'data aggregation',
]
CAT_B_LOW = [
    'heterogeneous', 'federated', 'distributed sources',
]
CAT_D = [
    'machine learning', 'artificial intelligence', 'deep learning', 'neural network',
    'autonomous', 'reinforcement learning', 'natural language processing',
    'large language model', 'LLM', 'AI',
]
CAT_E = [
    'data quality', 'data provenance', 'metadata', 'data governance', 'FAIR',
    'VAULTIS', 'trustworthy', 'data reliability', 'ontology', 'data catalog',
]
CAT_C = [
    'command and control', 'C2', 'decision support', 'situational awareness',
    'operational planning', 'battle management', 'military', 'operational',
]

# ── Adições manuais (idx = número da linha no CSV, base-1) ───────────────────
# Identificados por pesquisa textual no corpus após triagem automática.
# Incluídos por relevância temática justificada (ver docstring acima).
MANUAL_IDS = {
    # idx_csv : (file_id, justificativa_curta)
    # Schubert et al. 2018 — AI for Decision Support in C2 Systems
    # (23rd ICCRTS, posição 16 no CONF/JOURNAL do CSV → 23rd_ICCRTS_16)
    'schubert_2018': 'requisitos funcionais: o que a IA deve fazer com os dados (S2/S3, COP)',
    'abdelazez_2023_wargaming': 'Ministério da Defesa do Brasil + wargaming de COAs (Step 4 PPM)',
    'firth_2024': 'evidência empírica: data analytics como condição necessária para IA no planejamento',
    'turner_2024_challenging': 'evidência empírica: indisponibilidade de dados como principal barreira à IA',
}
# IDs físicos das adições manuais (obtidos da execução anterior do scoring)
MANUAL_FILE_IDS = [
    '23rd_ICCRTS_16',   # Schubert et al. 2018
    '28th_ICCRTS_04',   # Abdelazez et al. 2023 wargaming
    '29th_ICCRTS_27',   # Firth, Poole, Turner 2024
    '29th_ICCRTS_31',   # Turner et al. 2024 challenging case
]


def make_pattern(terms):
    return re.compile(
        r'\b(' + '|'.join(re.escape(t) for t in terms) + r')\b',
        re.IGNORECASE
    )


def hits(pattern, text):
    return set(m.lower() for m in pattern.findall(text))


def build_file_id(conf, year, seq_counter):
    """Constrói identificador físico legível a partir dos metadados do artigo."""
    c = conf.strip()
    m = re.search(r'(\d+)(st|nd|rd|th)', c, re.IGNORECASE)
    if m and ('iccrts' in c.lower() or 'command and control research' in c.lower()):
        ordinal = m.group(1) + m.group(2).lower()
        return f"{ordinal}_ICCRTS_{seq_counter[c]:02d}"
    if 'c2coe' in c.lower():
        return f"C2COE_{year}_{seq_counter[c]:02d}"
    if 'military operations research' in c.lower():
        return f"MOR_{year}_{seq_counter[c]:02d}"
    abbrev = re.sub(r'[^A-Za-z0-9]', '_', c)[:20].strip('_')
    return f"{abbrev}_{year}_{seq_counter[c]:02d}"


def convert_csv_if_needed():
    if not os.path.exists(CSV_IN):
        print("Convertendo xlsx para CSV via LibreOffice...")
        os.system(
            f"libreoffice --headless --convert-to csv '{XLSX_IN}' "
            f"--outdir /tmp/ 2>/dev/null"
        )


def compute_score(row):
    """Calcula todos os scores para um artigo."""
    title = row.get('TITLE', '')
    abst  = row.get('ABSTRACT', '')

    pat = compute_score._pat
    ah_t = hits(pat['A_HIGH'], title); ah_a = hits(pat['A_HIGH'], abst)
    al_t = hits(pat['A_LOW'],  title); al_a = hits(pat['A_LOW'],  abst)
    bh_t = hits(pat['B_HIGH'], title); bh_a = hits(pat['B_HIGH'], abst)
    bl_t = hits(pat['B_LOW'],  title); bl_a = hits(pat['B_LOW'],  abst)
    d_t  = hits(pat['D'],      title); d_a  = hits(pat['D'],      abst)
    e_t  = hits(pat['E'],      title); e_a  = hits(pat['E'],      abst)
    c_h  = hits(pat['C'],      title) | hits(pat['C'], abst)

    A = len(ah_t)*6 + len(ah_a)*3 + len(al_t)*3 + len(al_a)*1
    B = len(bh_t)*5 + len(bh_a)*3 + len(bl_t)*3 + len(bl_a)*1
    D = min(len(d_t)*3 + len(d_a)*2, D_CAP)
    E = len(e_t)*5  + len(e_a)*2
    bonus = BONUS_AB if (A + B) >= BONUS_THRESHOLD else 0
    comp  = A + B + D + E + bonus

    return {
        'score_A': A, 'score_B': B, 'score_D': D, 'score_E': E,
        'score_total': A + B + D + E,
        'score_comp': comp,
        'passes_auto': comp >= COMP_MIN and (A + B >= AB_MIN or E >= E_MIN),
        'hits_C': sorted(c_h),
        'detail': {
            'A_HIGH_title': sorted(ah_t), 'A_HIGH_abs': sorted(ah_a),
            'A_LOW_title':  sorted(al_t), 'A_LOW_abs':  sorted(al_a),
            'B_HIGH_title': sorted(bh_t), 'B_HIGH_abs': sorted(bh_a),
            'B_LOW_title':  sorted(bl_t), 'B_LOW_abs':  sorted(bl_a),
            'D_title': sorted(d_t), 'D_abs': sorted(d_a),
            'E_title': sorted(e_t), 'E_abs': sorted(e_a),
        },
    }


def run_scoring():
    convert_csv_if_needed()

    # Compila padrões uma única vez
    compute_score._pat = {
        'A_HIGH': make_pattern(CAT_A_HIGH),
        'A_LOW':  make_pattern(CAT_A_LOW),
        'B_HIGH': make_pattern(CAT_B_HIGH),
        'B_LOW':  make_pattern(CAT_B_LOW),
        'D':      make_pattern(CAT_D),
        'E':      make_pattern(CAT_E),
        'C':      make_pattern(CAT_C),
    }

    seq_counter = defaultdict(int)
    rows = []

    with open(CSV_IN, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, 1):
            conf   = row.get('CONFERENCE/JOURNAL', '')
            year   = row.get('YEAR', '')
            seq_counter[conf] += 1
            file_id = build_file_id(conf, year, seq_counter)

            sc = compute_score(row)
            if not sc['hits_C']:          # não passa pelo corte mínimo C
                continue

            rows.append({
                'idx':      i,
                'file_id':  file_id,
                'year':     year,
                'authors':  row.get('AUTHORS', ''),
                'title':    row.get('TITLE', ''),
                'conf':     conf,
                'source':   row.get('Source', ''),
                'manual':   file_id in MANUAL_FILE_IDS,
                **sc,
            })

    rows.sort(key=lambda r: (-r['score_comp'], -r['score_A'], -r['score_B']))
    return rows


def build_final_list(rows):
    """Retorna lista final: automáticos (passes_auto) + manuais."""
    auto    = [r for r in rows if r['passes_auto']]
    manual  = [r for r in rows if r['file_id'] in MANUAL_FILE_IDS
               and not r['passes_auto']]
    return auto, manual


def save_json(rows):
    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    print(f"  JSON salvo: {JSON_OUT}")


# ── Estilos xlsx ──────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F3864')
HDR_FONT   = Font(color='FFFFFF', bold=True, size=10)
MANUAL_FILL= PatternFill('solid', fgColor='D9E1F2')   # azul claro para manuais
AUTO_FILLS = {
    'high':   PatternFill('solid', fgColor='C6EFCE'),  # verde
    'mid':    PatternFill('solid', fgColor='FFEB9C'),  # amarelo
    'low':    PatternFill('solid', fgColor='FFCCCC'),  # vermelho claro
}
THIN = Side(style='thin', color='CCCCCC')
BRD  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def row_fill(r):
    if r['manual']:
        return MANUAL_FILL
    s = r['score_comp']
    if s >= 12: return AUTO_FILLS['high']
    if s >= 8:  return AUTO_FILLS['mid']
    return AUTO_FILLS['low']


def write_header(ws, cols):
    for ci, (name, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BRD
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A2'


def save_xlsx(auto, manual):
    wb = Workbook()

    # ── ABA 1: Candidatos finais ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Candidatos_33'

    cols1 = [
        ('Rank',         5), ('Seleção',      12), ('ID Físico',   22),
        ('Ano',          6), ('Autores',       32), ('Título',      62),
        ('Score Comp.',  10), ('A – Arq.',     9),  ('B – Integ.',  9),
        ('D – ML/IA',    9), ('E – Gov.',      9),  ('Bônus A+B',   9),
        ('Conferência',  36),
    ]
    write_header(ws1, cols1)
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(cols1))}1'

    all_final = [(r, 'Automático') for r in auto] + \
                [(r, 'Manual')     for r in manual]
    all_final.sort(key=lambda x: (-x[0]['score_comp'],
                                  0 if x[1]=='Automático' else 1))

    for rank, (r, sel) in enumerate(all_final, 1):
        bonus = BONUS_AB if (r['score_A'] + r['score_B']) >= BONUS_THRESHOLD else 0
        data = [
            rank, sel, r['file_id'], r['year'], r['authors'][:120],
            r['title'], r['score_comp'],
            r['score_A'], r['score_B'], r['score_D'], r['score_E'], bonus,
            r['conf'],
        ]
        fill = row_fill(r)
        for ci, val in enumerate(data, 1):
            cell = ws1.cell(row=rank+1, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=(ci >= 4))
            cell.border = BRD
            if fill: cell.fill = fill

    # ── ABA 2: Scoring detalhado ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Scoring_Detalhe')
    cols2 = [
        ('ID Físico', 22), ('Título', 55), ('Seleção', 12),
        ('Termos A_HIGH título', 35), ('Termos A_HIGH abs.', 35),
        ('Termos A_LOW título',  35), ('Termos A_LOW abs.',  35),
        ('Termos B_HIGH título', 35), ('Termos B_HIGH abs.', 35),
        ('Termos B_LOW título',  35), ('Termos B_LOW abs.',  35),
        ('Termos D título',      35), ('Termos D abs.',      35),
        ('Termos E título',      35), ('Termos E abs.',      35),
    ]
    write_header(ws2, cols2)

    for rank, (r, sel) in enumerate(all_final, 1):
        d = r['detail']
        data = [
            r['file_id'], r['title'], sel,
            ', '.join(d['A_HIGH_title']), ', '.join(d['A_HIGH_abs']),
            ', '.join(d['A_LOW_title']),  ', '.join(d['A_LOW_abs']),
            ', '.join(d['B_HIGH_title']), ', '.join(d['B_HIGH_abs']),
            ', '.join(d['B_LOW_title']),  ', '.join(d['B_LOW_abs']),
            ', '.join(d['D_title']),      ', '.join(d['D_abs']),
            ', '.join(d['E_title']),      ', '.join(d['E_abs']),
        ]
        fill = row_fill(r)
        for ci, val in enumerate(data, 1):
            cell = ws2.cell(row=rank+1, column=ci, value=val)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BRD
            if fill: cell.fill = fill

    # ── ABA 3: Metodologia ───────────────────────────────────────────────────
    ws3 = wb.create_sheet('Metodologia')
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 90

    secoes = [
        ('PARÂMETROS GERAIS', ''),
        ('D_CAP',
         f'{D_CAP} pts — pontuação máxima da categoria D (ML/IA) por artigo.\n'
         'Justificativa: sem o cap, artigos sobre AI genérica (realidade '
         'aumentada, UxVs, redes sociais) dominam o ranking por acumular '
         'muitas ocorrências de "AI" e "machine learning", mesmo sem tratar '
         'de arquitetura de dados. O valor 6 equivale a dois termos D no '
         'título — presença substantiva sem inflação por repetição.'),
        ('BÔNUS ARQUITETURAL',
         f'+ {BONUS_AB} pts quando (A + B) ≥ {BONUS_THRESHOLD}.\n'
         'Justificativa: artigos que combinam relevância arquitetural com '
         'outros critérios são os mais valiosos para a dissertação. O bônus '
         'eleva esses artigos sem introduzir novos termos. O threshold 3 '
         'corresponde ao peso mínimo que indica presença real do conceito '
         '(ex: um termo A_HIGH no abstract = 3 pts).'),
        ('', ''),
        ('REGRA DE CORTE AUTOMÁTICO', ''),
        ('Condição 1',
         f'score_comp ≥ {COMP_MIN}\n'
         'Justificativa: abaixo de 7, os artigos atingem o score '
         'exclusivamente via D=6 (cap), sem contribuição de A, B ou E. '
         'Tratam de AI/ML genérica, fora do escopo da dissertação.'),
        ('Condição 2',
         f'(A + B ≥ {AB_MIN})  OU  (E ≥ {E_MIN})\n'
         'Justificativa: score_comp ≥ 7 não é suficiente — artigos que '
         'atingem 7 via D=6 + 1 pt de A_LOW acidentalmente ainda são '
         'AI genérica. A condição adicional garante profundidade mínima '
         'verificável em dados (A+B ≥ 3) ou em governança (E ≥ 5).'),
        ('Resultado automático', f'{len(auto)} artigos'),
        ('', ''),
        ('ADIÇÕES MANUAIS', f'{len(manual)} artigos — elegibilidade por julgamento do pesquisador'),
    ]
    for m_id in MANUAL_FILE_IDS:
        row_m = next((r for r in manual if r['file_id'] == m_id), None)
        if row_m:
            secoes.append((m_id, row_m['title']))

    secoes += [
        ('', ''),
        ('CATEGORIAS — TERMOS', ''),
        ('A_HIGH (peso 6/3)', ', '.join(CAT_A_HIGH)),
        ('A_LOW  (peso 3/1)', ', '.join(CAT_A_LOW)),
        ('B_HIGH (peso 5/3)', ', '.join(CAT_B_HIGH)),
        ('B_LOW  (peso 3/1)', ', '.join(CAT_B_LOW)),
        ('D      (peso 3/2, cap=6)', ', '.join(CAT_D)),
        ('E      (peso 5/2)', ', '.join(CAT_E)),
        ('C      (corte mínimo, não pontua)', ', '.join(CAT_C)),
    ]

    hdr_fill3 = PatternFill('solid', fgColor='2E4057')
    for ri, (k, v) in enumerate(secoes, 1):
        ck = ws3.cell(row=ri, column=1, value=k)
        cv = ws3.cell(row=ri, column=2, value=v)
        ck.alignment = Alignment(vertical='top', wrap_text=True)
        cv.alignment = Alignment(vertical='top', wrap_text=True)
        if k and not v:   # cabeçalho de seção
            ck.font = Font(bold=True, color='FFFFFF')
            ck.fill = hdr_fill3
            cv.fill = hdr_fill3
        elif k:
            ck.font = Font(bold=True)
        ws3.row_dimensions[ri].height = max(15, min(80, len(str(v))//3))

    wb.save(XLSX_OUT)
    print(f"  XLSX salvo: {XLSX_OUT}")


# ── Execução ──────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("Executando scoring PRISMA revisado...")
    rows = run_scoring()
    auto, manual = build_final_list(rows)

    print(f"\n  Artigos que passaram pelo corte C: {len(rows)}")
    print(f"  Selecionados automaticamente    : {len(auto)}")
    print(f"  Adições manuais                 : {len(manual)}")
    print(f"  TOTAL candidatos                : {len(auto) + len(manual)}")

    all_final = sorted(
        [(r, 'Auto') for r in auto] + [(r, 'Manual') for r in manual],
        key=lambda x: (-x[0]['score_comp'], 0 if x[1]=='Auto' else 1)
    )

    print(f"\n{'Rk':<4} {'Sel':<7} {'Comp':>5} {'A':>5} {'B':>5} {'D':>4} {'E':>4} "
          f"| {'ID Físico':<24} {'Título'}")
    print('-' * 115)
    for rank, (r, sel) in enumerate(all_final, 1):
        print(f"{rank:<4} {sel:<7} {r['score_comp']:>5} {r['score_A']:>5} "
              f"{r['score_B']:>5} {r['score_D']:>4} {r['score_E']:>4} "
              f"| {r['file_id']:<24} {r['title'][:42]}")

    save_json(rows)
    save_xlsx(auto, manual)
    print("\nConcluído.")
